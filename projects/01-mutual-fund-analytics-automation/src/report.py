"""
Step 5 of the pipeline: turn scored_funds.csv into a formatted Excel
report - one sheet per category (ranked funds, top 3 highlighted) plus a
Summary sheet with the #1 pick from each category.

Input:  data/processed/scored_funds.csv
Output: output/fund_rankings.xlsx
"""

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

OUTPUT_PATH = "output/fund_rankings.xlsx"

# Columns shown on each category sheet, with their Excel number format.
DISPLAY_COLUMNS = {
    "category_rank": "0",
    "scheme_name": None,
    "amc": None,
    "total_aum_cr": "#,##0",
    "return_1y": "0.0%",
    "return_3y": "0.0%",
    "return_5y": "0.0%",
    "return_10y": "0.0%",
    "beta": "0.00",
    "sharpe": "0.00",
    "alpha": "0.0%",
    "consistency": "0.0%",
    "composite_score": "0.0%",
}

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOP_PICK_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")


def short_category_name(category: str) -> str:
    """Excel sheet names are capped at 31 chars - trim the common prefix/suffix."""
    return category.replace("Equity Scheme - ", "").replace(" Fund", "")


def style_sheet(worksheet, df: pd.DataFrame):
    """Apply header styling, column widths, number formats, and highlight
    the top-3 ranked rows on a single worksheet."""
    columns = list(df.columns)

    # Header row (row 1 in Excel, 1-indexed)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        is_top_pick = row["category_rank"] <= 3
        for col_idx, col_name in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            number_format = DISPLAY_COLUMNS.get(col_name)
            if number_format:
                cell.number_format = number_format
            if is_top_pick:
                cell.fill = TOP_PICK_FILL

    # Column widths - wider for names, narrower for numbers
    for col_idx, col_name in enumerate(columns, start=1):
        width = 38 if col_name in ("scheme_name", "amc") else 14
        worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = width


def build_summary(scored: pd.DataFrame) -> pd.DataFrame:
    """One row per category: the #1 ranked fund plus how many funds were eligible."""
    rows = []
    for category, group in scored.groupby("category"):
        top = group[group["category_rank"] == 1].iloc[0]
        rows.append({
            "category": short_category_name(category),
            "eligible_funds": len(group),
            "top_pick": top["scheme_name"],
            "amc": top["amc"],
            "composite_score": top["composite_score"],
            "sharpe": top["sharpe"],
            "alpha": top["alpha"],
            "consistency": top["consistency"],
        })
    return pd.DataFrame(rows)


def main():
    scored = pd.read_csv("data/processed/scored_funds.csv")

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        summary = build_summary(scored)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        for col_idx, col_name in enumerate(summary.columns, start=1):
            cell = writer.sheets["Summary"].cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            width = 38 if col_name in ("top_pick", "amc") else 16
            writer.sheets["Summary"].column_dimensions[cell.column_letter].width = width

        for category, group in scored.groupby("category"):
            sheet_name = short_category_name(category)
            display = group[list(DISPLAY_COLUMNS.keys())]
            display.to_excel(writer, sheet_name=sheet_name, index=False)
            style_sheet(writer.sheets[sheet_name], display)

    print(f"Saved report to {OUTPUT_PATH}")
    print(f"Sheets: Summary + {scored['category'].nunique()} categories")


if __name__ == "__main__":
    main()
