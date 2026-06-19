"""
Step 0 — extract raw price history from the original Excel workbook.

This is the only module that touches the .xlsx file. Every later pipeline
step works off the CSVs this script produces.
"""

import os

import openpyxl
import pandas as pd

DATA_RAW = os.path.join(os.path.dirname(__file__), "..", "data", "raw")
XLSX_PATH = os.path.join(DATA_RAW, "MASTER_2.0_clubChallenge.xlsx")
EQ_DEBT_CSV = os.path.join(DATA_RAW, "raw_eq_debt.csv")
GOLD_CSV = os.path.join(DATA_RAW, "raw_gold.csv")


def extract_eq_debt():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["RAW_EQ&DT"]
    rows = []
    for date, name, close in (
        (r[2], r[1], r[6]) for r in ws.iter_rows(min_row=2, values_only=True)
    ):
        # Rows with no Index Name are Gold data accidentally appended to this
        # sheet — Gold's real data lives in RAW_GOLD, so skip these here.
        if name is None or date is None:
            continue
        rows.append((date, name, close))
    df = pd.DataFrame(rows, columns=["Date", "Index Name", "Close"])
    df.to_csv(EQ_DEBT_CSV, index=False)
    return df


def extract_gold():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["RAW_GOLD"]
    header = [(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    date_col = header.index("Date")
    close_col = header.index("Close Price")
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        date, close = r[date_col], r[close_col]
        if date is None:
            continue
        rows.append((date, close))
    df = pd.DataFrame(rows, columns=["Date", "Close"])
    df.to_csv(GOLD_CSV, index=False)
    return df


def main():
    if os.path.exists(EQ_DEBT_CSV) and os.path.exists(GOLD_CSV):
        print("raw_eq_debt.csv and raw_gold.csv already exist, skipping extraction")
        return
    print("Extracting RAW_EQ&DT and RAW_GOLD from the Excel workbook...")
    eq_debt = extract_eq_debt()
    gold = extract_gold()
    print(f"  raw_eq_debt.csv: {len(eq_debt)} rows, {eq_debt['Index Name'].nunique()} indices")
    print(f"  raw_gold.csv: {len(gold)} rows")


if __name__ == "__main__":
    main()
