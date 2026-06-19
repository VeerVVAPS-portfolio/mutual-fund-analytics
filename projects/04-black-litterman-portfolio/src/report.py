"""
Step 7 of the pipeline: write the final Black-Litterman results to a
formatted Excel report — posterior returns, optimal allocation, stress
test/VaR, and an Assumptions sheet documenting every inherited quirk from
the original model (both "replica" and "corrected" figures side by side).

Output: output/black_litterman_report.xlsx
"""

import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "output")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "black_litterman_report.xlsx")

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

ASSUMPTIONS = [
    ("Equal-split equity market weights",
     "The 70.1% 'Indian Equities' market-cap bucket is split equally across the 8 equity "
     "sub-indices (8.76% each) rather than using each sector's individual market cap. "
     "Inherited from the original Excel, not redesigned."),
    ("Estimated market caps",
     "Asset-class market caps (Equities 475, G-Secs 195, Gold 2.5, REITs 5 - all in lakh-crore "
     "INR) are analyst estimates from the original model, not live AUM data."),
    ("25% per-asset weight cap",
     "Inferred from the original Solver output (G-Sec and Gold both land exactly on 25%), "
     "not stated explicitly anywhere in the workbook."),
    ("Covariance matrix window",
     "Per-asset return/vol statistics use the full 2015-2025 history, but the covariance "
     "matrix (which everything downstream depends on) is computed only over the period "
     "REITS has listed data (Aug 2019 onward), using population covariance. This mixes two "
     "different lookback windows - an inherited inconsistency from the original Excel."),
    ("Monthly covariance used as if annual",
     "The original model never built a separate annualized covariance matrix - the same "
     "MONTHLY covariance matrix is used throughout reverse optimization, Black-Litterman "
     "blending, and the optimizer, while annual returns and the annual risk-free rate are "
     "combined into the same formulas. This is the root cause of the Sharpe Ratio / VaR "
     "unit mismatch below. The 'Replica' columns reproduce this exactly; 'Corrected' columns "
     "use the annualized covariance matrix consistently throughout instead."),
    ("Sharpe Ratio / VaR bug",
     "Found while reverse-engineering the formulas: Portfolio Volatility is computed from "
     "the monthly covariance matrix while Portfolio Return is annualized, so the published "
     "Sharpe Ratio (1.64) divides an annual return by a monthly volatility. The VaR formula "
     "compounds the same error by dividing an already-monthly volatility by sqrt(12) again. "
     "See the Stress Test & VaR sheet for both the replica and corrected figures."),
    ("Omega matrix uses cross-terms",
     "Standard Black-Litterman assumes views are uncorrelated (diagonal Omega). The original "
     "model's Omega is the full P*Sigma*P^T including off-diagonal cross-terms between views "
     "- replicated exactly in Replica mode; Corrected mode diagonalizes Omega and optionally "
     "scales it by each view's stated confidence (Idzorek convention), which the original "
     "model's confidence inputs never actually fed into."),
]


def style_header(ws, n_cols):
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def main(results: dict):
    """
    `results` is a dict produced by main.py's pipeline run, containing the
    replica and corrected outputs for each stage. See main.py for its shape.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    posterior_df = pd.DataFrame({
        "Replica (matches Excel)": results["posterior_replica"],
        "Corrected (consistent annual units)": results["posterior_corrected"],
    })

    weights_replica = pd.Series(results["optimizer_replica"]["weights"])
    weights_corrected = pd.Series(results["optimizer_corrected"]["weights"])
    allocation_df = pd.DataFrame({
        "Replica weight": weights_replica,
        "Corrected weight": weights_corrected,
    })

    metrics_df = pd.DataFrame({
        "Replica (matches Excel)": [
            results["optimizer_replica"]["return"],
            results["optimizer_replica"]["volatility"],
            results["optimizer_replica"]["sharpe"],
            results["stressed_return"],
            results["var_replica"],
        ],
        "Corrected": [
            results["optimizer_corrected"]["return"],
            results["optimizer_corrected"]["volatility"],
            results["optimizer_corrected"]["sharpe"],
            results["stressed_return"],
            results["var_corrected"],
        ],
    }, index=["Portfolio Return", "Portfolio Volatility", "Sharpe Ratio", "Stressed Return (-20% equity shock)", "95% Monthly VaR"])

    assumptions_df = pd.DataFrame(ASSUMPTIONS, columns=["Assumption / Finding", "Detail"])

    with pd.ExcelWriter(OUTPUT_PATH, engine="openpyxl") as writer:
        posterior_df.to_excel(writer, sheet_name="Posterior Returns")
        style_header(writer.sheets["Posterior Returns"], posterior_df.shape[1] + 1)

        allocation_df.to_excel(writer, sheet_name="Optimal Allocation")
        style_header(writer.sheets["Optimal Allocation"], allocation_df.shape[1] + 1)

        metrics_df.to_excel(writer, sheet_name="Stress Test & VaR")
        style_header(writer.sheets["Stress Test & VaR"], metrics_df.shape[1] + 1)

        assumptions_df.to_excel(writer, sheet_name="Assumptions", index=False)
        style_header(writer.sheets["Assumptions"], assumptions_df.shape[1])
        writer.sheets["Assumptions"].column_dimensions["A"].width = 32
        writer.sheets["Assumptions"].column_dimensions["B"].width = 100

    print(f"Saved report to {OUTPUT_PATH}")


if __name__ == "__main__":
    raise SystemExit("Run via main.py - report.main() needs the full pipeline results dict.")
